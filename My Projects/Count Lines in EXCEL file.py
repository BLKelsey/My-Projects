from openpyxl import load_workbook          # Import function to open Excel (.xlsx) files

print("Enter a filename: ")                 # Ask the user for the Excel file name
filename = input()                          # Store the filename the user types

workbook = load_workbook(filename)          # Open the Excel file
sheet = workbook.active                     # Select the first (active) worksheet

row_count = 0                               # Start counting rows at zero

for row in sheet.iter_rows(values_only=True):  # Go through each row in the sheet.  iter_rows() literally means:“Go through the worksheet one row at a time.”
    row_count += 1                             # Add 1 for each row found

print("Number of rows: ", row_count)        # Display the total number of rows
