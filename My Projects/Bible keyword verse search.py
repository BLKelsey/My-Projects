from openpyxl import load_workbook
import random
import os

# Show which file is being used (helps catch wrong-file issues)
print(os.path.abspath("RefAndScripture.xlsx"))

# Load the Excel file
workbook = load_workbook("RefAndScripture.xlsx")
sheet = workbook.active

# Bible book order (used for sorting)
bible_books = [
    "Genesis", "Exodus", "Leviticus", "Numbers", "Deuteronomy",
    "Joshua", "Judges", "Ruth",
    "1 Samuel", "2 Samuel", "1 Kings", "2 Kings",
    "1 Chronicles", "2 Chronicles", "Ezra", "Nehemiah", "Esther",
    "Job", "Psalm", "Proverbs", "Ecclesiastes", "Song of Solomon",
    "Isaiah", "Jeremiah", "Lamentations", "Ezekiel", "Daniel",
    "Hosea", "Joel", "Amos", "Obadiah", "Jonah", "Micah",
    "Nahum", "Habakkuk", "Zephaniah", "Haggai", "Zechariah", "Malachi",
    "Matthew", "Mark", "Luke", "John", "Acts",
    "Romans", "1 Corinthians", "2 Corinthians", "Galatians",
    "Ephesians", "Philippians", "Colossians",
    "1 Thessalonians", "2 Thessalonians",
    "1 Timothy", "2 Timothy", "Titus", "Philemon",
    "Hebrews", "James", "1 Peter", "2 Peter",
    "1 John", "2 John", "3 John", "Jude", "Revelation"
]

# Ask the user for a keyword or phrase
keyword = input("Enter a keyword or phrase to search for: ").strip().lower()

matches = []

# Search through the verses
for row in sheet.iter_rows(min_row=2, values_only=True):
    reference, verse = row

    if verse is None:
        continue

    text = verse.lower()

    # Phrase search (contains spaces)
    if " " in keyword:
        if keyword in text:
            matches.append((reference, verse))
    else:
        # Single word search (exact word)
        words = text.split()
        if keyword in words:
            matches.append((reference, verse))

# Limit to 10 random verses
if len(matches) > 10:
    matches = random.sample(matches, 10)

# Helper function to sort by Bible book order
def get_book_order(reference):
    book_name = reference.rsplit(" ", 1)[0]
    return bible_books.index(book_name)

# Sort results by Bible book order
matches.sort(key=lambda item: get_book_order(item[0]))

# Display results
print("\nSearch results:\n")
for reference, verse in matches:
    print(f"{reference} - {verse}")

print(f"\nTotal verses shown: {len(matches)}")
